from openpyxl import load_workbook

from ..utils import ensure_unmerged
from openpyxl.utils import range_boundaries


class BangKeWriter:
    def __init__(self, path, sheet_name):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb[sheet_name]

    def write_orders(self, orders, start_row):
        self._prepare_data_area(start_row)
        row = start_row

        for order in orders:
            merge = order["merge"]
            start = row
            end = row + merge - 1

            if merge > 1:
                self.ws.insert_rows(start + 1, merge - 1)

            # --- Write MERGED columns (TOP-LEFT ONLY) ---
            for col, val in order["base"].items():
                ensure_unmerged(self.ws, start, col)

                cell = self.ws[f"{col}{start}"]

                # Long number → TEXT
                if isinstance(val, int) and len(str(val)) >= 11:
                    cell.value = str(val)
                    cell.number_format = "@"

                # Column N → NUMBER + accounting
                elif col == "N":
                    try:
                        cell.value = float(val)
                        cell.number_format = "#,##0"  # '#,##0.00' nếu cần
                    except:
                        cell.value = val

                else:
                    cell.value = val

            # --- Write conts & cars ---
            for i in range(merge):
                if i < len(order["conts"]):
                    ensure_unmerged(self.ws, start + i, "I")
                    self.ws[f"I{start + i}"].value = order["conts"][i]

                if i < len(order["cars"]):
                    ensure_unmerged(self.ws, start + i, "J")
                    self.ws[f"J{start + i}"].value = order["cars"][i]

            # --- ALWAYS write FORMULA for column O ---
            ensure_unmerged(self.ws, start, "O")
            o_cell = self.ws[f"O{start}"]
            o_cell.value = f"=N{start}/(K{start}+L{start})"

            o_cell.number_format = "#,##0.00"

            # --- Merge AFTER writing ---
            if merge > 1:
                for col in list("ABCDEFGH") + ["K", "L", "M", "N", "O"]:
                    self.ws.merge_cells(f"{col}{start}:{col}{end}")

            row = end + 1

        # self.wb.save(self.path)
        return row

    def write_phu_phi_row(self, row, p, q, t_formula=None, order_start_row=None):
        if p is not None:
            ensure_unmerged(self.ws, row, "P")
            cell_p = self.ws[f"P{row}"]
            cell_p.value = p
            cell_p.alignment = cell_p.alignment.copy(wrap_text=True)

        if q is not None:
            ensure_unmerged(self.ws, row, "Q")
            cell_q = self.ws[f"Q{row}"]
            cell_q.value = q
            cell_q.alignment = cell_q.alignment.copy(wrap_text=True)

        # --- xử lý công thức ---
        if t_formula and order_start_row:
            formula = str(t_formula)

            # thay K, L theo HÀNG ĐẦU TIÊN CỦA ĐƠN
            formula = formula.replace("K", f"K{order_start_row}")
            formula = formula.replace("L", f"L{order_start_row}")

            ensure_unmerged(self.ws, row, "T")
            self.ws[f"T{row}"].value = f"={formula}"
        # nếu t_formula trống → không làm gì

    def write_order_total(self, order_start_row, order_end_row):
        if order_end_row < order_start_row:
            return

        # merge cột X
        if order_end_row > order_start_row:
            self.ws.merge_cells(
                f"X{order_start_row}:X{order_end_row}"
            )

        # gắn công thức SUM
        ensure_unmerged(self.ws, order_start_row, "X")
        self.ws[f"X{order_start_row}"].value = (
            f"=SUM(W{order_start_row}:W{order_end_row})"
        )

    def _prepare_data_area(self, start_row, end_col="Z"):
        """
        Unmerge tất cả merged-range nằm TỪ start_row TRỞ XUỐNG
        KHÔNG ĐỤNG HEADER
        """
        to_unmerge = []

        for rng in self.ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))

            # chỉ unmerge vùng data
            if min_row >= start_row:
                to_unmerge.append(str(rng))

        for rng in to_unmerge:
            self.ws.unmerge_cells(rng)