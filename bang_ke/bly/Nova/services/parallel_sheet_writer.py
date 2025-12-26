from openpyxl.cell.cell import MergedCell
from ..utils import ensure_unmerged


class ParallelSheetWriter:
    def __init__(self, bang_ke_ws):
        self.ws = bang_ke_ws

    def _get_source_cell(self, ws, row, col_letter):
        cell = ws[f"{col_letter}{row}"]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )
        return cell

    def _copy_cell(self, src_cell, dst_row, dst_col):
        # 🔥 UNMERGE TRƯỚC
        ensure_unmerged(self.ws, dst_row, dst_col)

        dst_cell = self.ws[f"{dst_col}{dst_row}"]
        dst_cell.value = src_cell.value

        if src_cell.has_style:
            dst_cell.font = src_cell.font.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.alignment = src_cell.alignment.copy()
            dst_cell.number_format = src_cell.number_format

    def write_parallel(self, theo_doi_ws, orders, mapping, start_row=6):
        current_row = start_row

        for order in orders:
            src_row = order.row_idx
            dst_row = current_row

            for m in mapping:
                src_col = m["src"]
                dst_col = m["dst"]

                src_cell = self._get_source_cell(
                    theo_doi_ws, src_row, src_col
                )

                self._copy_cell(src_cell, dst_row, dst_col)

            current_row += 1

        return current_row
