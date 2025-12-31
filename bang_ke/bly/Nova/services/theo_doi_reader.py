from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ..models import OrderModel
import re


class TheoDoiReader:
    def __init__(self, path):
        self.path = path

    def _parse_month(self, text):
        if not isinstance(text, str):
            return None
        match = re.search(r"THÁNG\s*(\d{1,2})", text.upper())
        if match:
            return int(match.group(1))
        return None

    def read_nova_by_month(self, sheet_name: str, month: int):
        print(f"=== READ THEO DOI {sheet_name} ===")
        print(f"Target month: {month}")

        wb = load_workbook(self.path, data_only=True)
        ws = wb[sheet_name]

        orders = []
        current_month = None

        for r in range(1, ws.max_row + 1):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r))[0]

            # --- Detect MONTH ROW ---
            detected_month = None
            for cell in row_cells:
                detected_month = self._parse_month(cell.value)
                if detected_month:
                    break

            if detected_month:
                current_month = detected_month
                print(f"[MONTH] Row {r}: {cell.value} → month={detected_month}")
                continue

            # --- Skip rows not in selected month ---
            if current_month != month:
                continue

            row_data = {}
            has_data = False

            for idx, cell in enumerate(row_cells, start=1):
                col_letter = get_column_letter(idx)  # ✅ KHÔNG PHỤ THUỘC cell
                row_data[col_letter] = cell.value
                if cell.value not in (None, ""):
                    has_data = True

            if has_data:
                orders.append(OrderModel(r, row_data))
                print(f"[ORDER] Row {r} added")

        print(f"=== TOTAL NHAP ORDERS FOUND: {len(orders)} ===")
        return orders

    # def read_nova_nhap_by_month(self, month, sheet_name="NOVA NHẬP"):
    #     print("=== READ THEO DOI NOVA NHẬP ===")
    #     print(f"Target month: {month}")
    #
    #     wb = load_workbook(self.path, data_only=True)
    #     ws = wb[sheet_name]
    #
    #     orders = []
    #     current_month = None
    #
    #     for r in range(1, ws.max_row + 1):
    #         row_cells = list(ws.iter_rows(min_row=r, max_row=r))[0]
    #
    #         # --- Detect MONTH ROW ---
    #         detected_month = None
    #         for cell in row_cells:
    #             detected_month = self._parse_month(cell.value)
    #             if detected_month:
    #                 break
    #
    #         if detected_month:
    #             current_month = detected_month
    #             print(f"[MONTH] Row {r}: {cell.value} → month={detected_month}")
    #             continue
    #
    #         # --- Skip rows not in selected month ---
    #         if current_month != month:
    #             continue
    #
    #         row_data = {}
    #         has_data = False
    #
    #         for idx, cell in enumerate(row_cells, start=1):
    #             col_letter = get_column_letter(idx)  # ✅ KHÔNG PHỤ THUỘC cell
    #             row_data[col_letter] = cell.value
    #             if cell.value not in (None, ""):
    #                 has_data = True
    #
    #         if has_data:
    #             orders.append(OrderModel(r, row_data))
    #             print(f"[ORDER] Row {r} added")
    #
    #     print(f"=== TOTAL NHAP ORDERS FOUND: {len(orders)} ===")
    #     return orders
    #
    # def read_nova_xuat_by_month(self, month, sheet_name="NOVA XUẤT"):
    #     print("=== READ THEO DOI NOVA XUẤT ===")
    #     print(f"Target month: {month}")
    #
    #     wb = load_workbook(self.path, data_only=True)
    #     ws = wb[sheet_name]
    #
    #     orders = []
    #     current_month = None
    #
    #     for r in range(1, ws.max_row + 1):
    #         row_cells = list(ws.iter_rows(min_row=r, max_row=r))[0]
    #
    #         # --- Detect MONTH ROW ---
    #         detected_month = None
    #         for cell in row_cells:
    #             detected_month = self._parse_month(cell.value)
    #             if detected_month:
    #                 break
    #
    #         if detected_month:
    #             current_month = detected_month
    #             print(f"[MONTH] Row {r}: {cell.value} → month={detected_month}")
    #             continue
    #
    #         # --- Skip rows not in selected month ---
    #         if current_month != month:
    #             continue
    #
    #         row_data = {}
    #         has_data = False
    #
    #         for idx, cell in enumerate(row_cells, start=1):
    #             col_letter = get_column_letter(idx)  # ✅ KHÔNG PHỤ THUỘC cell
    #             row_data[col_letter] = cell.value
    #             if cell.value not in (None, ""):
    #                 has_data = True
    #
    #         if has_data:
    #             orders.append(OrderModel(r, row_data))
    #             print(f"[ORDER] Row {r} added")
    #
    #     print(f"=== TOTAL XUAT ORDERS FOUND: {len(orders)} ===")
    #     return orders
