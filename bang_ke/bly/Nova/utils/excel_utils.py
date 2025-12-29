from openpyxl.utils import get_column_letter

def ensure_unmerged(ws, row, col):
    """
    Unmerge merged-range chứa (row, col).
    Nếu openpyxl lỗi nội bộ → gỡ range thủ công.
    """

    if isinstance(col, int):
        col_letter = get_column_letter(col)
    else:
        col_letter = col

    coord = f"{col_letter}{row}"

    for rng in list(ws.merged_cells.ranges):
        if coord in rng:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                # QUAN TRỌNG: gỡ merged range thủ công
                ws.merged_cells.ranges.remove(rng)
            break
