import openpyxl


def load_plate_mapping(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "bien_so_xe" not in wb.sheetnames:
        raise Exception("Không tìm thấy sheet 'bien_so_xe' trong mapping.xlsx")

    ws = wb["bien_so_xe"]
    plate_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        sheet_name = str(row[0]).strip()
        plate = str(row[1]).strip()
        plate_map[sheet_name] = plate

    return plate_map


def load_location_mapping(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "dia_diem" not in wb.sheetnames:
        raise Exception("Không tìm thấy sheet 'dia_diem' trong mapping.xlsx")

    ws = wb["dia_diem"]
    location_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1] or not row[2]:
            continue

        token = str(row[0]).strip().upper()
        full_name = str(row[1]).strip()
        province = str(row[2]).strip()
        location_map[token] = (full_name, province)

    return location_map
