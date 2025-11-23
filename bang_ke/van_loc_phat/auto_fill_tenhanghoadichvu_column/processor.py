import openpyxl
import re


def split_tokens(raw):
    if not raw:
        return []

    raw = raw.upper().replace("/", "-").replace(",", "-")
    parts = [p.strip() for p in raw.split("-") if p.strip()]
    return parts


def convert_tokens(tokens, location_map, log):
    """Chuyển đổi list tokens thành chuỗi đã chuẩn hóa
    Nếu token không có mapping => ghi log và return None để báo sheet đó bỏ qua dòng
    """
    mapped = []

    for t in tokens:
        if t not in location_map:
            log.append(f"Token '{t}' không có trong mapping → bỏ qua dòng.")
            return None  # báo hiệu bỏ qua dòng

        full_name, province = location_map[t]
        mapped.append((full_name, province))

    # nhóm theo province
    province_groups = {}
    order = []

    for name, province in mapped:
        if province not in province_groups:
            province_groups[province] = []
            order.append(province)
        province_groups[province].append(name)

    # ghép chuỗi theo thứ tự tỉnh
    result_parts = []
    for prov in order:
        names = province_groups[prov]
        joined_names = " - ".join(names)
        result_parts.append(f"{joined_names}, {prov}")

    return " - ".join(result_parts)


def process_bang_ke(input_path, mapping_path, output_path):
    from mapping_loader import load_plate_mapping, load_location_mapping

    log = []
    log.append("=== BẮT ĐẦU XỬ LÝ ===")

    plate_map = load_plate_mapping(mapping_path)
    location_map = load_location_mapping(mapping_path)

    wb = openpyxl.load_workbook(input_path)

    for sheet_name in wb.sheetnames:
        if sheet_name not in plate_map:
            log.append(f"[{sheet_name}] Không có biển số → bỏ qua sheet.")
            continue

        ws = wb[sheet_name]
        plate = plate_map[sheet_name]

        log.append(f"\n--- Sheet: {sheet_name} (Biển số: {plate}) ---")

        for row in ws.iter_rows(min_row=2):
            cell_d = row[3].value  # cột D
            cell_g = row[6]  # cột G

            if not cell_d:
                continue

            tokens = split_tokens(cell_d)
            converted = convert_tokens(tokens, location_map, log)

            if converted is None:
                cell_g.value = ""  # bỏ trống dòng
                continue

            final_text = (
                f"Biển số xe {plate} - Cước vận chuyển từ Vsip II, Bình Dương đi "
                f"{converted} (Địa Chỉ Cũ)"
            )

            cell_g.value = final_text

    wb.save(output_path)
    log.append("\n=== HOÀN THÀNH ===")
    return "\n".join(log)
